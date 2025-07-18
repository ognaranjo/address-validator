import os
import time
from urllib import response
import pandas as pd
from usps_api import validate_usps
from chatgpt_api import analyze_with_chatgpt
from google_places_api import search_google_places, validate_google_address
from chatgpt_organize_address import organize_address_with_chatgpt
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import PatternFill
from utils import build_full_google_address


# ===== CONFIGURATION =====
INPUT_FILE = "addresses.xlsx"
PROCESSED_FILE = "processed_addresses.csv"
OUTPUT_FILE = "validated_addresses.xlsx"  # Only used if SAVE_AS_NEW_FILE is True
SAVE_AS_NEW_FILE = True  # Set to False to save as new tab in original file
NEW_TAB_NAME = "Validated"


usps_api_in_use = False  # Set to False if you want to use Google Places API only

google_places_cache = {}

# ===== LOAD DATA =====
# Check if processed file exists
if os.path.exists(PROCESSED_FILE):
    df = pd.read_csv(PROCESSED_FILE, dtype={'tx_town': str,'cd_state': str,'cd_zip': str,'curr_adr': str,'grp_cnt': str,'ad_strt_nbr': str,'ad_strt_nme': str,'ad_line2': str,'ad_line3': str,'apt': str,'street_name_adjusted_for_matching': str,'new_town': str,'new_state': str,'new_zip': str,'new_addr_nbr': str,'new_address': str,'new_apt': str,'error': str,'status': str,'notes': str}, low_memory=False)
else:
    df = pd.read_excel(INPUT_FILE, dtype={'TX_TOWN': str,'CD STATE': str,'CD_ZIP': str,'Curr Adr': str,'Grp cnt': str,'AD_STRT NBR': str,'AD_STRT_NME': str,'AD Line2': str,'AD Line3': str,'Apt': str,'Street Name Adjusted for Matching': str,'NEW_TOWN': str,'NEW STATE': str,'NEW ZIP': str,'NEW ADDR NBR': str,'NEW_ADDRESS': str,'NEW APT': str,'Error': str,'status': str,'notes': str})

# Normalize column names
df.columns = df.columns.str.strip().str.replace(' ', '_').str.lower()

# Reset index to track Excel row number
# df.reset_index(inplace=True)
# df.rename(columns={'index': 'pandas_index'}, inplace=True)
# df['original_excel_row_number'] = df['pandas_index'] + 2

# Add result columns
# df['new_town'] = ''
# df['new_state'] = ''
# df['new_zip'] = ''
# df['new_addr_nbr'] = ''
# df['new_address'] = ''
# df['new_apt'] = ''
# df['addr_ref'] = ''
# df['status'] = ''
# df['notes'] = ''

# List of your target columns (converted to lowercase)
columns_to_print = [
    'tx_town',
    'cd_state',
    'cd_zip',
    'curr_adr',
    'grp_cnt',
    'ad_strt_nbr',
    'ad_strt_nme',
    'ad_line2',
    'ad_line3',
    'apt'
]

#columns_to_convert = ['tx_town',     'cd_state',    'cd_zip',    'curr_adr',    'grp_cnt',    'ad_strt_nbr',    'ad_strt_nme',    'ad_line2',    'ad_line3',    'apt',    'street_name_adjusted_for_matching',    'new_town',    'new_state',    'new_zip',    'new_addr_nbr',    'new_address',    'new_apt',    'error',     'status',     'notes'  ]
#for col in columns_to_convert:
#    if col in df.columns:
#        # Convert to string, then replace any nan/NaN/NAN with empty string
#        df[col] = df[col].astype(str).replace(r'(?i)\bnan\b', '', regex=True)


# ===== PROCESS ROWS =====
for index, row in df.iterrows():
    error = row.get('error', None)
    status = row.get('status', '')

    # Skip if Error is blank or null
    if pd.isna(error) or str(error).strip() == '':
        continue

    # Skip if Status is not blank or null 
    if pd.notna(status) and str(status).strip() != '':
        continue
    
    # Skip if all address fields are blank
    if all(pd.isna(row.get(col, '')) or str(row.get(col, '')).strip() == '' for col in ['ad_strt_nbr', 'ad_strt_nme', 'ad_line2', 'ad_line3']):
       df.loc[index, 'notes'] = 'All address fields blank'
       df.loc[index, 'status'] = 'NotFound'
       
       # error_col_letter = 'R'  # Replace with actual column letter for Error

       # df[f'{error_col_letter}{index}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

       # Save after each decision to prevent data loss
       df.to_csv(PROCESSED_FILE, index=False)

       # Save your dataframe first if it does not exist
       #if not os.path.exists(OUTPUT_FILE):
       #    df.to_csv(OUTPUT_FILE, index=False)

       # Highlight Error column in yellow (if saving via openpyxl later)
       # You will need to apply formatting after saving with pandas:
       # Save first, then reopen with openpyxl:
       # wb = openpyxl.load_workbook(OUTPUT_FILE)
       # ws = wb.active
       # excel_row = index + 2  # Adjust for header
       
       # ws[f'{error_col_letter}{excel_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

       # wb.save(OUTPUT_FILE)
       # wb.close()
    
       continue


    bad_values = ("deceased", "unknown", "homeless")
    skip_cols = ['ad_strt_nbr', 'ad_strt_nme', 'ad_line2', 'ad_line3']

    def contains_bad_value(val):
        val_str = str(val).strip().lower()
        return any(bad in val_str for bad in bad_values)

    # If any field contains "Deceased" or "UNKNOWN" (case-insensitive)
    if any(contains_bad_value(row.get(col, '')) for col in skip_cols):
        df.loc[index, 'notes'] = 'Address contains DECEASED or UNKNOWN'
        df.loc[index, 'status'] = 'NotFound'
        df.to_csv(PROCESSED_FILE, index=False)
        continue



    # STEP 1. Organize address with ChatGPT
    organized = organize_address_with_chatgpt(row)
    if not organized:
        df.loc[index, 'status'] = 'NoMatch'
        df.loc[index, 'notes'] = 'ChatGPT unable to parse address'
        # Save after each decision to prevent data loss
        df.to_csv(PROCESSED_FILE, index=False)
        continue


    print("************ ROW START ************")
    print(f"Processing row {index + 2}")
    
    for col in columns_to_print:
        value = row.get(col, '')
        print(f"{col.upper()}: {value}")    


    print(f"PLACE NAME: {organized['place_name']}")

    google_validation_queue = []    

    # STEP 2. Validate with USPS

    street_number = organized.get('street_number') or ''
    street_name = organized.get('street_name') or ''

    if not street_number or street_number.strip().lower() == 'n/a' or street_number.strip().lower() == '':
       full_street_address = ''
    else:
        full_street_address = f"{street_number} {street_name}".strip()


    if usps_api_in_use:
        usps_result = validate_usps(
            full_street_address, 
            organized['city'], 
            organized['state'], 
            organized['zip'],
            organized.get('apt', '')
        )
        if not usps_result['valid']:
            if usps_result.get("error") == "Rate limit exceeded":
                usps_api_in_use = False
    else:
        usps_result = {
            "valid": False,
            "corrected_address": None,
            "error": "USPS API not not in use"
        }

    if not usps_api_in_use:
       if full_street_address and full_street_address.strip():
           google_validation_queue.append(build_full_google_address(full_street_address, organized['city'], organized['state'], organized['zip'], organized.get('apt', '')))
       for alt_zipcode in organized['nearby_zipcodes']:
           alt_zipcode_full = alt_zipcode.strip()
           if full_street_address and full_street_address.strip():
              google_validation_queue.append(build_full_google_address(full_street_address, organized['state'], alt_zipcode_full, organized.get('apt', '')))   
       for alt_address in organized['alternative_addresses']:
            alt_address_full = alt_address.strip()

            # First, try with the current zip if it exists
            if organized.get('zip'):
               if alt_address_full and alt_address_full.strip():
                  google_validation_queue.append(build_full_google_address(alt_address_full, '', organized['state'],  organized['zip'], organized.get('apt', '')))

            if organized.get('nearby_zipcodes'):
                for alt_zipcode in organized['nearby_zipcodes']:
                    alt_zipcode_full = alt_zipcode.strip()
                    if alt_address_full and alt_address_full.strip():
                       google_validation_queue.append(build_full_google_address(alt_address_full, '', organized['state'],  alt_zipcode_full, organized.get('apt', '')))


    # If USPS validation fails, try nearby zipcodes
    if not usps_result['valid'] and organized.get('nearby_zipcodes') and usps_api_in_use:
        if full_street_address and full_street_address.strip():
           google_validation_queue.append(build_full_google_address(full_street_address, organized['city'], organized['state'], organized['zip'], organized.get('apt', '')))
        for alt_zipcode in organized['nearby_zipcodes']:
            alt_zipcode_full = alt_zipcode.strip()
            time.sleep(0.5)
            alt_usps_result = validate_usps(
                full_street_address, 
                '',  # no city, 
                organized['state'], 
                alt_zipcode_full,
                organized.get('apt', '')
            )
            if alt_usps_result['valid']:
                usps_result = alt_usps_result
                # You may optionally note that this is an alternative address match
                df.loc[index, 'notes'] = f"Used alternative zipcode: {alt_zipcode_full}"
                break  # Stop after first valid alternative
            else:
                # If USPS validation succeeded, no need to queue for Google validation
                if full_street_address and full_street_address.strip():
                    google_validation_queue.append(build_full_google_address(full_street_address, '', organized['state'], alt_zipcode_full, organized.get('apt', '')))

    # If USPS validation fails, try alternative addresses with current zip and nearby zipcodes
    if not usps_result['valid'] and organized.get('alternative_addresses') and usps_api_in_use:
        for alt_address in organized['alternative_addresses']:
            alt_address_full = alt_address.strip()

            # First, try with the current zip if it exists
            if organized.get('zip'):
                time.sleep(0.5)  # Avoid hitting rate limits
                alt_usps_result = validate_usps(
                    alt_address_full,
                    '',  # no city
                    organized['state'],
                    organized['zip'],
                    organized.get('apt', '')
                )
                if alt_usps_result['valid']:
                    usps_result = alt_usps_result
                    df.loc[index, 'notes'] = f" Used alternative address with current zip: {organized['zip']}"
                    break  # Stop after first valid alternative
                else:
                    # If USPS validation succeeded, no need to queue for Google validation
                    if alt_address_full and alt_address_full.strip():
                        google_validation_queue.append(build_full_google_address(alt_address_full, '', organized['state'],  organized['zip'], organized.get('apt', '')))


            # Then, try with each nearby zipcode
            if organized.get('nearby_zipcodes'):
                for alt_zipcode in organized['nearby_zipcodes']:
                    alt_zipcode_full = alt_zipcode.strip()
                    time.sleep(0.5)
                    alt_usps_result = validate_usps(
                        alt_address_full,
                        '',  # no city
                        organized['state'],
                        alt_zipcode_full,
                        organized.get('apt', '')
                    )
                    if alt_usps_result['valid']:
                        usps_result = alt_usps_result
                        df.loc[index, 'notes'] = f" Used alternative address with nearby zip: {alt_zipcode_full}"
                        break  # Stop after first valid alternative
                    else:
                        # If USPS validation succeeded, no need to queue for Google validation
                        if alt_address_full and alt_address_full.strip():
                            google_validation_queue.append(build_full_google_address(alt_address_full, '', organized['state'],  alt_zipcode_full, organized.get('apt', '')))

            # Break outer loop if a valid result was found in either of the above
            if usps_result['valid']:
                break



    if usps_result['valid']:
        # Update NEW_* columns with USPS validated data
        df.loc[index, 'new_town'] = usps_result['standardized_address'].get('city', organized['city'])
        df.loc[index, 'new_state'] =  usps_result['standardized_address'].get('state', organized['state'])
        df.loc[index, 'new_zip'] = usps_result['standardized_address'].get('zip5', organized['zip'])
        df.loc[index, 'new_addr_nbr'] = usps_result['standardized_address'].get('street_number', organized['street_number'])   
        df.loc[index, 'new_address'] = usps_result['standardized_address'].get('street_name', organized['street_name']) 
        df.loc[index, 'new_apt'] = usps_result['standardized_address'].get('apt', organized['apt']) 
        df.loc[index, 'addr_ref'] = organized['ref']
        df.loc[index, 'status'] = 'Valid'
        # Save after each decision to prevent data loss
        df.to_csv(PROCESSED_FILE, index=False)
    else:

        for address in google_validation_queue:
            google_result = validate_google_address(address)
            if google_result['valid']:
                usps_result = google_result

                # Update NEW_* columns with USPS validated data
                df.loc[index, 'new_town'] = google_result['parsed_address'].get('city', organized['city'])
                df.loc[index, 'new_state'] =  google_result['parsed_address'].get('state', organized['state'])
                df.loc[index, 'new_zip'] = google_result['parsed_address'].get('zip5', '')
                df.loc[index, 'new_addr_nbr'] = google_result['parsed_address'].get('street_number', '')   
                df.loc[index, 'new_address'] = google_result['parsed_address'].get('street_name', '') 
                df.loc[index, 'new_apt'] = google_result['parsed_address'].get('apt', '') 
                df.loc[index, 'addr_ref'] = organized['ref']
                df.loc[index, 'status'] = 'Valid'
                df.loc[index, 'notes'] = 'ValidatedByGoogle: ' + google_result['corrected_address'] 
                df.loc[index, 'google_maps_url'] = google_result['google_maps_url']
                # Save after each decision to prevent data loss
                df.to_csv(PROCESSED_FILE, index=False)
                break

        if not usps_result['valid'] and organized['place_name'] and organized['place_name'].strip().lower() not in ['n/a', 'na']:
            google_result = search_google_places(organized['place_name'], google_places_cache, organized['state'])
            if google_result['found']:
     
                # Update NEW_* columns with USPS validated data
                df.loc[index, 'new_town'] = google_result['parsed_address'].get('city', organized['city'])
                df.loc[index, 'new_state'] =  google_result['parsed_address'].get('state', organized['state'])
                df.loc[index, 'new_zip'] = google_result['parsed_address'].get('zip5', '')
                df.loc[index, 'new_addr_nbr'] = google_result['parsed_address'].get('street_number', '')   
                df.loc[index, 'new_address'] = google_result['parsed_address'].get('street_name', '') 
                df.loc[index, 'addr_ref'] = organized['ref']
                df.loc[index, 'status'] = 'Valid'
                df.loc[index, 'notes'] = 'PlaceNameFound: ' + google_result['place_id'] + ' - ' + google_result['name']
                # Save after each decision to prevent data loss
                df.to_csv(PROCESSED_FILE, index=False)

            else:
                df.loc[index, 'status'] = 'NoMatch'
                df.loc[index, 'notes'] = 'Place name not found in Google Places'
                df.to_csv(PROCESSED_FILE, index=False)
        else:
            # If no valid address was found, mark as NoMatch
            if not usps_result['valid']:
                df.loc[index, 'status'] = 'NoMatch'
                df.loc[index, 'notes'] = 'No valid address found'   
                df.to_csv(PROCESSED_FILE, index=False)
    print("************* ROW END *************\n")

# ===== OUTPUT RESULTS =====
if SAVE_AS_NEW_FILE:
    df.to_csv(PROCESSED_FILE, index=False)
    print(f"✅ Results saved to new file: {PROCESSED_FILE}")
else:
    book = load_workbook(PROCESSED_FILE)
    writer = pd.ExcelWriter(PROCESSED_FILE, engine='openpyxl')
    writer.book = book
    df.to_excel(writer, sheet_name=NEW_TAB_NAME, index=False)
    writer.save()
    writer.close()
    print(f"✅ Results saved as new tab '{NEW_TAB_NAME}' in file: {PROCESSED_FILE}")
